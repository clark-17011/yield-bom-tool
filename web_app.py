import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
import unicodedata
import io
import matplotlib.pyplot as plt
import matplotlib
import platform
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ==========================================
# 1. 頁面設定
# ==========================================
st.set_page_config(page_title="Yield & BOM Tool", layout="wide", page_icon="📊")

# 初始化 Session
if 'exp_yield_open' not in st.session_state: st.session_state['exp_yield_open'] = True
if 'exp_bom_open' not in st.session_state: st.session_state['exp_bom_open'] = True

# ==========================================
# CSS 優化 (v4.2 樣式保持不變)
# ==========================================
st.markdown("""
    <style>
        header[data-testid="stHeader"] {
            background-color: transparent !important;
            border-bottom: none !important;
            pointer-events: none !important;
            z-index: 100 !important;
        }
        [data-testid="stDecoration"], [data-testid="stToolbar"] { display: none !important; }
        
        button[data-testid="baseButton-header"], button[data-testid="stSidebarCollapsedControl"] {
            display: block !important;
            visibility: visible !important;
            pointer-events: auto !important;
            color: #444 !important;
            background-color: rgba(255, 255, 255, 0.6) !important;
            border-radius: 50%;
            margin-top: 0.5rem;
        }

        .block-container {
            padding-top: 3rem !important;
            padding-bottom: 2rem !important;
            padding-left: 3rem !important;
            padding-right: 3rem !important;
        }
        
        .streamlit-expanderHeader {
            background-color: #f8f9fa;
            border-radius: 4px;
            border: 1px solid #eee;
        }

        section[data-testid="stSidebar"] .block-container {
            padding-left: 0rem !important;
            padding-right: 0rem !important;
            padding-top: 4rem !important;
        }
        section[data-testid="stSidebar"] h1, 
        section[data-testid="stSidebar"] .stMarkdown,
        section[data-testid="stSidebar"] hr,
        section[data-testid="stSidebar"] .stCaption {
            padding-left: 1.5rem !important;
            padding-right: 1.5rem !important;
        }

        section[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label > div:first-child {
            display: none !important;
        }
        section[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label {
            width: 100% !important;
            padding: 15px 20px 15px 24px !important;
            margin: 0px !important;
            border: none !important;
            display: flex !important;
            font-size: 16px !important;
            transition: background-color 0.2s;
        }
        section[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label:hover {
            background-color: rgba(0, 0, 0, 0.05) !important;
        }
        section[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label:has(input:checked) {
            background-color: transparent !important;
            color: #000000 !important;
            font-weight: 700 !important;
            border-left: 5px solid #ff4b4b !important;
        }
        footer {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 0. 基礎設定與工具
# ==========================================

def configure_chart_font():
    system_name = platform.system()
    if system_name == "Windows":
        plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei', 'SimHei', 'Arial']
    elif system_name == "Darwin": 
        plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'PingFang TC', 'Heiti TC']
    else:
        plt.rcParams['font.sans-serif'] = ['WenQuanYi Zen Hei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False 

configure_chart_font()

def normalize_str(x) -> str:
    if x is None: return ""
    s = str(x)
    s = unicodedata.normalize("NFKC", s)
    s = s.strip()
    # [修正] 允許中文 (\u4e00-\u9fa5) 通過，只移除其他特殊符號
    # 這樣 "錫膏" 就不會變成 "" (空字串)
    s = re.sub(r"[^A-Za-z0-9\u4e00-\u9fa5]", "", s)
    return s.lower()

def parse_search_config(raw_text: str, is_space_or_mode: bool):
    if not raw_text.strip(): return []
    raw_text = raw_text.replace("，", ",")
    segments = [s.strip() for s in raw_text.split(',') if s.strip()]
    configs = []
    for seg in segments:
        if seg.startswith('[') and seg.endswith(']'):
            content = seg[1:-1].strip()
            sub_terms = [t.strip() for t in content.split() if t.strip()]
            if sub_terms:
                configs.append({'display': seg, 'terms': sub_terms})
            continue
        parts = [p.strip() for p in seg.split() if p.strip()]
        if not parts: continue
        if is_space_or_mode:
            for p in parts: configs.append({'display': p, 'terms': [p]})
        else:
            if len(parts) > 1:
                display_name = f"[{' '.join(parts)}]"
                configs.append({'display': display_name, 'terms': parts})
            else:
                configs.append({'display': parts[0], 'terms': [parts[0]]})
    
    seen = set()
    unique = []
    for c in configs:
        if c['display'] not in seen:
            unique.append(c)
            seen.add(c['display'])
    return unique

# ==========================================
# 1. Yield Report 邏輯
# ==========================================

@st.cache_data(ttl=3600, show_spinner=False)
def load_yield_files(uploaded_files):
    raw_data = [] 
    row_texts = [] 
    total_sheets = 0
    
    for file in uploaded_files:
        name = file.name
        label = "CPC" if "CPC" in name.upper() else ("HELE" if "HELE" in name.upper() else name)
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows: continue
                    header = rows[0]
                    data_rows = rows[1:]
                    if not data_rows: continue
                    raw_data.append({"label": label, "sheet": sheet_name, "header": header, "rows": data_rows})
                    current_sheet_texts = []
                    for row in data_rows:
                        joined = "".join([str(c) if c is not None else "" for c in row])
                        current_sheet_texts.append(normalize_str(joined))
                    row_texts.append({"label": label, "sheet": sheet_name, "texts": current_sheet_texts})
                    total_sheets += 1
                except: pass
            wb.close()
        except Exception as e: print(f"Error loading {label}: {e}")
            
    return raw_data, row_texts, total_sheets

def execute_yield_search(raw_data, row_texts, configs):
    if not configs: return pd.DataFrame(), set()
    prepared_configs = []
    for cfg in configs:
        norm_terms = [normalize_str(t) for t in cfg['terms'] if t.strip()]
        if norm_terms: prepared_configs.append({'display': cfg['display'], 'terms': norm_terms})

    all_rows_data = []
    found_display_names = set()
    
    for idx, sheet_info in enumerate(row_texts):
        label = sheet_info["label"]
        sheet_name = sheet_info["sheet"]
        sheet_norm_texts = sheet_info["texts"]
        header = raw_data[idx]["header"]
        all_rows = raw_data[idx]["rows"]
        
        unique_header = []
        seen_counts = {}
        for col in header:
            c_str = str(col).strip() if col is not None else ""
            if not c_str: c_str = "Unnamed"
            if c_str in seen_counts:
                seen_counts[c_str] += 1
                new_name = f"{c_str}.{seen_counts[c_str]}"
            else:
                seen_counts[c_str] = 0
                new_name = c_str
            unique_header.append(new_name)
            
        for row_idx, row_str in enumerate(sheet_norm_texts):
            for cfg in prepared_configs:
                is_match = True
                for term in cfg['terms']:
                    if term not in row_str:
                        is_match = False; break
                if is_match:
                    found_display_names.add(cfg['display'])
                    original_row = all_rows[row_idx]
                    row_dict = {"MatchedKeyword": cfg['display'], "SourceLabel": label, "SheetName": sheet_name}
                    for h_idx, col_name in enumerate(unique_header):
                        val = original_row[h_idx] if h_idx < len(original_row) else None
                        row_dict[col_name] = val
                    all_rows_data.append(row_dict)

    if all_rows_data:
        df_result = pd.DataFrame(all_rows_data)
        cols = list(df_result.columns)
        sys_cols = ['MatchedKeyword', 'SourceLabel', 'SheetName']
        other_cols = [c for c in cols if c not in sys_cols]
        df_result = df_result[sys_cols + other_cols]
    else:
        df_result = pd.DataFrame()

    all_targets = set(c['display'] for c in prepared_configs)
    missing = all_targets - found_display_names
    return df_result, missing

# ==========================================
# 2. BOM Tool 邏輯
# ==========================================
PCB_VENDOR_MAP = {"P": "PRV", "S": "SCC", "U": "旭德", "H": "AKM", "D": "科佳"}
PCB_FINISH_MAP = {"G": "化金", "N": "鎳鈀金", "P": "OSP"}
BASE_OUTPUT_ORDER = ["MPN", "Device Name", "ASIC (簡化 BOM)", "Sensor (簡化 BOM)", "PCB 供應商", "錫膏", "PCB 簡化 BOM", "金屬殼", "電容 / 電組 / 電感", "磁珠", "防水膜 / 金屬網", "Coating"]

def format_value(val):
    if pd.isna(val) or val is None: return ""
    val_str = str(val).replace("\n", " ").replace("\r", " ")
    return " ".join(val_str.split())

@st.cache_data(ttl=3600, show_spinner=False)
def load_bom_files(uploaded_files):
    raw_data = [] 
    row_texts = [] 
    for file in uploaded_files:
        name = file.name
        label = "CPC" if "CPC" in name.upper() else ("HELE" if "HELE" in name.upper() else name)
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows: continue
                    header = rows[0]
                    data_rows = rows[1:]
                    if not data_rows: continue
                    raw_data.append({"label": label, "sheet": sheet_name, "header": header, "rows": data_rows})
                    current_sheet_texts = []
                    for row in data_rows:
                        joined = "".join([str(c) if c is not None else "" for c in row])
                        current_sheet_texts.append(normalize_str(joined))
                    row_texts.append({"label": label, "sheet": sheet_name, "texts": current_sheet_texts})
                except: pass
            wb.close()
        except Exception as e: print(f"Error loading {label}: {e}")
    return raw_data, row_texts

def parse_pcb_details(green_bom):
    details = []
    if pd.isna(green_bom) or not green_bom: return details
    tokens = re.split(r'[\s\n]+', str(green_bom).strip())
    for token in tokens:
        token = token.strip()
        if len(token) >= 10:
            v_code = token[-4]
            f_code = token[-2]
            vendor = PCB_VENDOR_MAP.get(v_code, None)
            finish = PCB_FINISH_MAP.get(f_code, None)
            details.append({'code': token, 'vendor': vendor, 'finish': finish})
    return details

def get_col_by_keyword(header, keyword, exclude=None):
    target_key = normalize_str(keyword)
    for idx, col in enumerate(header):
        col_str = str(col)
        col_key = normalize_str(col_str)
        if target_key in col_key:
            if exclude and exclude in col_str: continue
            return idx
    return None

def execute_bom_search(raw_data, row_texts, terms_raw):
    export_list = []
    found_terms = set()
    norm_terms_map = {t: normalize_str(t) for t in terms_raw}
    for term in terms_raw:
        n_term = norm_terms_map[term]
        if not n_term: continue
        for idx, sheet_info in enumerate(row_texts):
            sheet_norm_texts = sheet_info["texts"]
            matched_row_idx = -1
            for r_idx, row_str in enumerate(sheet_norm_texts):
                if n_term in row_str:
                    matched_row_idx = r_idx; break
            if matched_row_idx != -1:
                found_terms.add(term)
                header = raw_data[idx]["header"]
                row = raw_data[idx]["rows"][matched_row_idx]
                label = sheet_info["label"]
                row_data = extract_bom_data(header, row, label, term)
                export_list.append(row_data)
    missing_terms = set(terms_raw) - found_terms
    return export_list, missing_terms

def extract_bom_data(header, row, label, term):
    row_data = {"Search Term": term, "Source File": label}
    def get_val(idx): return row[idx] if idx is not None and idx < len(row) else None

    pcb_green_idx = get_col_by_keyword(header, "PCB BOM (Green)")
    val_green = get_val(pcb_green_idx)
    pcb_details = parse_pcb_details(val_green)
    vendors = sorted(list(set([d['vendor'] for d in pcb_details if d['vendor']])))
    
    pcb_simple_idx = get_col_by_keyword(header, "PCB 簡化 BOM")
    if not pcb_simple_idx: pcb_simple_idx = get_col_by_keyword(header, "PCB 簡化BOM")
    val_simple = format_value(get_val(pcb_simple_idx))
    if not vendors and val_simple:
        for code, name in PCB_VENDOR_MAP.items():
            if name in val_simple: vendors.append(name)
    vendors_str = " / ".join(vendors) if vendors else ""

    LAYOUT_PLAN = [
        ("MPN", "normal", ["MPN"]), ("Device Name", "normal", ["Device Name"]),
        ("ASIC (簡化 BOM)", "merge", ["ASIC", "ASIC 簡化BOM"]), ("Sensor (簡化 BOM)", "merge", ["Sensor ID", "Sensor 簡化BOM"]),
        ("PCB 供應商", "value", vendors_str), ("錫膏", "normal", ["錫膏"]),
        ("PCB 簡化 BOM", "normal", ["PCB 簡化BOM"]), ("PCB List", "pcb_list", pcb_details),
        ("金屬殼", "normal", ["金屬殼 BOM (Blue)"]), ("電容 / 電組 / 電感", "normal", ["Indigo"]),
        ("磁珠", "normal", ["磁珠"]), ("防水膜 / 金屬網", "normal", ["防水膜"]), ("Coating", "normal", ["Coating BOM (Black)"]),
    ]

    for label_text, type_, args in LAYOUT_PLAN:
        final_val = ""
        if type_ == "normal":
            col_key = args[0]
            exclude = "簡化" if "簡化" not in label_text and "簡化" not in col_key else None
            if "Indigo" in col_key: exclude = None
            idx = get_col_by_keyword(header, col_key, exclude=exclude)
            val = format_value(get_val(idx))
            if val: final_val = val
        elif type_ == "merge":
            main_key, sub_key = args
            idx_main = get_col_by_keyword(header, main_key, exclude="簡化")
            idx_sub = get_col_by_keyword(header, sub_key)
            val_main = format_value(get_val(idx_main))
            val_sub = format_value(get_val(idx_sub))
            if val_main and val_sub: final_val = f"{val_main} ({val_sub})" if val_main != val_sub else val_main
            elif val_main: final_val = val_main
            elif val_sub: final_val = val_sub
        elif type_ == "value": final_val = str(args)
        elif type_ == "pcb_list":
            for i, item in enumerate(args, 1):
                code = item['code']; finish = item['finish']
                display = code + (f" ({finish})" if finish else "")
                row_data[f"PCB {i}"] = display
            continue
        if final_val: row_data[label_text] = final_val
    return row_data

# ==========================================
# 3. Streamlit UI 主程式
# ==========================================

# 1. 側邊欄 (功能選單)
with st.sidebar:
    st.title("功能選單")
    app_mode = st.radio("Mode", ["Yield Analysis", "BOM Search"], label_visibility="collapsed")
    st.divider()
    st.caption("v4.5 Chinese Fixed")

# --- Callbacks ---
def close_yield_expander():
    st.session_state['exp_yield_open'] = False

def close_bom_expander():
    st.session_state['exp_bom_open'] = False

# --- Page 1: Yield Analysis ---
if app_mode == "Yield Analysis":
    
    with st.expander("Yield 檔案載入與搜尋", expanded=st.session_state['exp_yield_open']):
        c1, c2 = st.columns([1, 1.2])
        with c1:
            yield_files = st.file_uploader("1. 上傳 Report", type=['xlsx', 'xls'], accept_multiple_files=True, key="yu")
            st.caption("自動辨識: HELE / CPC")
        with c2:
            raw_search = st.text_area("2. 關鍵字", height=68, placeholder="[Device A], [Device B]...", help="空格=AND, 逗號=OR")
            cc1, cc2 = st.columns([1, 1])
            chk_space = cc1.checkbox("空白=OR", value=False)
            btn_search_yield = cc2.button("搜尋", type="primary", key="by", use_container_width=True, on_click=close_yield_expander)

    yield_raw, yield_rows = [], []
    has_data = False
    
    if yield_files:
        current_files_key = ",".join([f"{f.name}-{f.size}" for f in yield_files])
        if st.session_state.get('last_yield_key') != current_files_key:
            with st.spinner("Loading..."): 
                yield_raw, yield_rows, sht_cnt = load_yield_files(yield_files)
            st.session_state['yield_cache'] = (yield_raw, yield_rows, sht_cnt)
            st.session_state['last_yield_key'] = current_files_key
            st.toast(f"已載入 {len(yield_files)} 檔, {sht_cnt} Sheets", icon="✅")
        else:
            yield_raw, yield_rows, sht_cnt = st.session_state.get('yield_cache', ([], [], 0))
        has_data = True
    elif 'yield_cache' in st.session_state:
        yield_raw, yield_rows, sht_cnt = st.session_state['yield_cache']
        if yield_raw:
            st.info(f"ℹ️ 使用已快取的資料 (共 {sht_cnt} Sheets)")
            has_data = True

    if btn_search_yield and has_data:
        if not raw_search.strip(): st.warning("請輸入關鍵字")
        else:
            cfgs = parse_search_config(raw_search, chk_space)
            with st.spinner("Searching..."): df_res, mis = execute_yield_search(yield_raw, yield_rows, cfgs)
            if mis: st.error(f"Missing: {', '.join(mis)}")
            if not df_res.empty:
                for c in df_res.columns:
                    if "DATE" in c.upper() or "TIME" in c.upper() or "日期" in c:
                        try: df_res[c] = pd.to_datetime(df_res[c], errors='coerce')
                        except: pass
                st.session_state['yield_result'] = df_res
            else:
                st.info("無符合資料")
                st.session_state['yield_result'] = pd.DataFrame()
    elif btn_search_yield and not has_data:
        st.error("請先上傳檔案")

    if 'yield_result' in st.session_state and not st.session_state['yield_result'].empty:
        full_df = st.session_state['yield_result']
        
        with st.expander("資料篩選", expanded=True):
            fc1, fc2, fc3 = st.columns(3)
            sel_kw = fc1.selectbox("對象", ["(全部)"] + sorted(list(full_df['MatchedKeyword'].unique())))
            sel_sht = fc2.selectbox("Sheet", ["(全部)"] + sorted(list(full_df['SheetName'].unique())))
            dcols = [c for c in full_df.columns if pd.api.types.is_datetime64_any_dtype(full_df[c])]
            sel_date = fc3.selectbox("時間", ["(無)"] + dcols)
            
            fdf = full_df.copy()
            if sel_kw != "(全部)": fdf = fdf[fdf['MatchedKeyword'] == sel_kw]
            if sel_sht != "(全部)": fdf = fdf[fdf['SheetName'] == sel_sht]
            if sel_date != "(無)":
                dmin, dmax = fdf[sel_date].min(), fdf[sel_date].max()
                if pd.notnull(dmin) and pd.notnull(dmax):
                    sd, ed = st.date_input("區間", value=(dmin, dmax), min_value=dmin, max_value=dmax)
                    fdf = fdf[(fdf[sel_date].dt.date >= sd) & (fdf[sel_date].dt.date <= ed)]
            
            if sel_kw != "(全部)" or sel_sht != "(全部)": fdf = fdf.dropna(axis=1, how='all')
            curr = fdf.columns.tolist()
            sys = ['MatchedKeyword', 'SourceLabel', 'SheetName']
            fsys = [c for c in sys if c in curr]
            fdata = [c for c in curr if c not in sys]
            fdf = fdf[fsys + fdata]
            st.caption(f"Count: {len(fdf)}")

        t1, t2 = st.tabs(["詳細數據", "統計圖表"])
        
        with t1:
            st.dataframe(fdf, use_container_width=True)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                font = Font(bold=True, color="FFFFFF")
                fill = PatternFill("solid", fgColor="4472C4")
                for s in fdf['SheetName'].unique():
                    sub = fdf[fdf['SheetName']==s].dropna(axis=1, how='all')
                    sname = str(s)[:30]
                    sub.to_excel(writer, sheet_name=sname, index=False)
                    ws = writer.sheets[sname]
                    for cell in ws[1]:
                        cell.font = font; cell.fill = fill; cell.alignment = Alignment(horizontal='center')
                    for col in ws.columns:
                        mlen = 0
                        cl = get_column_letter(col[0].column)
                        for cell in col:
                            try: mlen = max(mlen, len(str(cell.value)))
                            except: pass
                        ws.column_dimensions[cl].width = min((mlen+2)*1.1, 60)
            st.download_button("下載 Excel", data=buf.getvalue(), file_name="yield_result.xlsx")

        with t2:
            ncols = fdf.select_dtypes(include=['number']).columns.tolist()
            acols = fdf.columns.tolist()
            dcols_sort = [c for c in acols if pd.api.types.is_datetime64_any_dtype(fdf[c]) or "DATE" in c.upper() or "TIME" in c.upper() or "日期" in c]
            sort_opts = ["(X軸預設值)"] + dcols_sort + [c for c in acols if c not in dcols_sort]

            def agg_fn(df, x, y, f):
                if f == "計數 (Count)": return df[x].value_counts().sort_index().reset_index(name='Count').rename(columns={'index':x, x:x})
                d = df.copy(); d[y] = pd.to_numeric(d[y], errors='coerce').fillna(0)
                if f == "平均 (Mean)": return d.groupby(x)[y].mean().reset_index()
                elif f == "加總 (Sum)": return d.groupby(x)[y].sum().reset_index()
                elif f == "最大值 (Max)": return d.groupby(x)[y].max().reset_index()
                return d

            def sort_fn(pdf, x, sc, odf):
                if sc == "(X軸預設值)" or sc not in odf.columns:
                    try: return pdf.sort_values(by=x)
                    except: return pdf
                m = pd.merge(pdf, odf.groupby(x)[sc].min().reset_index(), on=x, how='left')
                return m.sort_values(by=sc)

            mc1, mc2 = st.columns([1, 1])
            mode = mc1.radio("模式", ["單軸", "雙軸"], horizontal=True, label_visibility="collapsed")
            scol = mc2.selectbox("排序依據", sort_opts)
            st.divider()

            fig = None
            if mode == "單軸":
                pc1, pc2, pc3, pc4 = st.columns(4)
                ptype = pc1.selectbox("類型", ["Bar","Line","Pie","Scatter"])
                pxax = pc2.selectbox("X", acols)
                pfunc = pc4.selectbox("算", ["平均 (Mean)","加總 (Sum)","最大值 (Max)","計數 (Count)"])
                pyax = "(Count)" if pfunc=="計數 (Count)" else pc3.selectbox("Y", ncols)
                
                if st.button("繪圖", key="draw1", use_container_width=True):
                    pdf = fdf[pxax].value_counts().reset_index(name='Count') if pfunc=="計數 (Count)" else agg_fn(fdf, pxax, pyax, pfunc)
                    if pfunc=="計數 (Count)": pdf.columns = [pxax, 'Count']; yname='Count'
                    else: yname=pyax
                    pdf = sort_fn(pdf, pxax, scol, fdf)
                    tt = f"{pfunc} - {yname} by {pxax}"
                    if ptype=="Bar": fig=px.bar(pdf, x=pxax, y=yname, title=tt, text_auto='.2s')
                    elif ptype=="Line": fig=px.line(pdf, x=pxax, y=yname, markers=True, title=tt)
                    elif ptype=="Pie": fig=px.pie(pdf, names=pxax, values=yname, title=tt)
                    elif ptype=="Scatter": fig=px.scatter(pdf, x=pxax, y=yname, title=tt)

            else: # Combo
                cc1, cc2, cc3 = st.columns(3)
                cxax = cc1.selectbox("X", acols)
                with cc2:
                    st.caption("左軸")
                    cy1t = st.selectbox("圖", ["Bar","Line"], key="cy1t")
                    cf1 = st.selectbox("算", ["平均 (Mean)","加總 (Sum)","最大值 (Max)","計數 (Count)"], key="cf1")
                    cy1 = "Count" if cf1=="計數 (Count)" else st.selectbox("值", ncols, key="cy1")
                with cc3:
                    st.caption("右軸")
                    cy2t = st.selectbox("圖", ["Line","Bar"], key="cy2t")
                    cf2 = st.selectbox("算", ["平均 (Mean)","加總 (Sum)","最大值 (Max)","計數 (Count)"], key="cf2")
                    cy2 = "Count" if cf2=="計數 (Count)" else st.selectbox("值", ncols, index=1 if len(ncols)>1 else 0, key="cy2")

                if st.button("繪圖", key="draw2", use_container_width=True):
                    d1 = agg_fn(fdf, cxax, cy1, cf1)
                    d2 = agg_fn(fdf, cxax, cy2, cf2)
                    n1 = "Count" if cf1=="計數 (Count)" else cy1
                    n2 = "Count" if cf2=="計數 (Count)" else cy2
                    if n1==n2: 
                        d1.rename(columns={n1:n1+"_L"}, inplace=True); n1+="_L"
                        d2.rename(columns={n2:n2+"_R"}, inplace=True); n2+="_R"
                    
                    m = pd.merge(d1, d2, on=cxax, how='outer').fillna(0)
                    m = sort_fn(m, cxax, scol, fdf)
                    
                    fig = make_subplots(specs=[[{"secondary_y": True}]])
                    t1 = go.Bar(x=m[cxax], y=m[n1], name=f"{n1}({cf1})", marker_color='#007AFF', opacity=0.7) if cy1t=="Bar" else go.Scatter(x=m[cxax], y=m[n1], name=f"{n1}({cf1})", mode='lines+markers', line=dict(color='#007AFF'))
                    fig.add_trace(t1, secondary_y=False)
                    t2 = go.Bar(x=m[cxax], y=m[n2], name=f"{n2}({cf2})", marker_color='#FF453A', opacity=0.7) if cy2t=="Bar" else go.Scatter(x=m[cxax], y=m[n2], name=f"{n2}({cf2})", mode='lines+markers', line=dict(color='#FF453A', width=3))
                    fig.add_trace(t2, secondary_y=True)
                    fig.update_layout(title=f"{cf1} vs {cf2}", hovermode="x unified", legend=dict(orientation="h", y=1.02))
                    fig.update_yaxes(title=n1, secondary_y=False); fig.update_yaxes(title=n2, secondary_y=True)

            if fig: st.plotly_chart(fig, use_container_width=True)


# --- Page 2: BOM Search ---
elif app_mode == "BOM Search":
    
    with st.expander("BOM 檔案載入與搜尋", expanded=st.session_state['exp_bom_open']):
        c1, c2 = st.columns([1, 1.2])
        with c1:
            bom_files = st.file_uploader("1. 上傳 BOM", type=['xlsx', 'xls'], accept_multiple_files=True, key="bu")
        with c2:
            bom_input = st.text_area("2. 料號", height=68)
            cc1, cc2 = st.columns([1, 1])
            chk_b_space = cc1.checkbox("空白分隔", key="bs")
            btn_bom = cc2.button("比對", type="primary", use_container_width=True, on_click=close_bom_expander)

    # === [關鍵修改] 智慧快取邏輯 (BOM) ===
    bom_raw, bom_rows = [], []
    has_bom_data = False
    
    if bom_files:
        current_bom_key = ",".join([f"{f.name}-{f.size}" for f in bom_files])
        if st.session_state.get('last_bom_key') != current_bom_key:
            with st.spinner("Indexing..."): 
                bom_raw, bom_rows = load_bom_files(bom_files)
            st.session_state['bom_cache'] = (bom_raw, bom_rows)
            st.session_state['last_bom_key'] = current_bom_key
            st.toast(f"BOM Index Ready: {len(bom_files)} files", icon="✅")
        else:
            bom_raw, bom_rows = st.session_state.get('bom_cache', ([], []))
        has_bom_data = True
    elif 'bom_cache' in st.session_state:
        bom_raw, bom_rows = st.session_state['bom_cache']
        if bom_raw:
            st.info("ℹ️ 使用已快取的 BOM 資料")
            has_bom_data = True

    if btn_bom and has_bom_data:
        if not bom_input.strip(): st.warning("請輸入料號")
        else:
            sep = r'[,\n\r\t\s，;]+' if chk_b_space else r'[,\n\r\t，;]+'
            terms = list(dict.fromkeys([t.strip() for t in re.split(sep, bom_input) if t.strip()]))
            st.caption(f"Searching {len(terms)} items...")
            
            with st.spinner("Searching..."): res, missing = execute_bom_search(bom_raw, bom_rows, terms)
            
            if missing: st.error(f"Missing: {', '.join(missing)}")
            else: st.success("All Found!")
            
            if res:
                dfb = pd.DataFrame(res)
                mpcb = 0
                for r in res:
                    for k in r.keys():
                        if k.startswith("PCB ") and k[4:].isdigit(): mpcb = max(mpcb, int(k[4:]))
                
                hdrs = ["Search Term", "Source File"]
                bo = list(BASE_OUTPUT_ORDER)
                try: idx = bo.index("PCB 簡化 BOM") + 1
                except: idx = len(bo)
                fcols = hdrs + bo[:idx] + [f"PCB {i}" for i in range(1, mpcb+1)] + bo[idx:]
                dfb = dfb.reindex(columns=fcols)
                
                st.dataframe(dfb, use_container_width=True)
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine='openpyxl') as w:
                    dfb.to_excel(w, index=False, sheet_name='Res')
                    for c in w.sheets['Res'].columns: 
                        w.sheets['Res'].column_dimensions[get_column_letter(c[0].column)].width = 20
                st.download_button("下載 BOM 結果", data=buf.getvalue(), file_name="BOM_Result.xlsx")
    elif btn_bom and not has_bom_data:
        st.error("請先上傳 BOM 檔案")
